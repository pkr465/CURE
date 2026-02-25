"""
CURE — Codebase Update & Refactor Engine
Codebase Patch Agent

Self-contained patch analysis agent.  Analyses patches (unified diffs)
against source files to identify issues **introduced by the patch**.

The agent is fully independent of :class:`CodebaseLLMAgent`.  It:
  1. Parses the unified diff and applies hunks to reconstruct the patched file.
  2. Extracts only the code regions around the hunks (not the whole file).
  3. Gathers 4-layer context from the real codebase (header context,
     context validation, call-stack analysis, constraints).
  4. Sends the hunk-scoped code + context to the LLM using
     ``PATCH_REVIEW_PROMPT`` and calls ``LLMTools.llm_call()`` directly.
  5. Parses the LLM response, post-filters to hunk ranges, and writes
     findings to a ``patch_<filename>`` tab in ``detailed_code_review.xlsx``.
"""

import logging
import os
import re
import tempfile
import shutil
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

# ---------------------------------------------------------------------------
# Graceful imports
# ---------------------------------------------------------------------------

# Patch Review prompt (dedicated, NOT the general analysis prompt)
try:
    from prompts.patch_review_prompt import PATCH_REVIEW_PROMPT
    PATCH_PROMPT_AVAILABLE = True
except ImportError:
    PATCH_REVIEW_PROMPT = None
    PATCH_PROMPT_AVAILABLE = False

# Context layer: header resolution
try:
    from agents.context.header_context_builder import HeaderContextBuilder
    HEADER_CTX_AVAILABLE = True
except ImportError:
    HeaderContextBuilder = None
    HEADER_CTX_AVAILABLE = False

# Context layer: context validation (false-positive reduction)
try:
    from agents.context.context_validator import ContextValidator
    CTX_VALIDATOR_AVAILABLE = True
except ImportError:
    ContextValidator = None
    CTX_VALIDATOR_AVAILABLE = False

# Context layer: static call-stack tracing
try:
    from agents.context.static_call_stack_analyzer import StaticCallStackAnalyzer
    CALL_STACK_AVAILABLE = True
except ImportError:
    StaticCallStackAnalyzer = None
    CALL_STACK_AVAILABLE = False

# Constraint loader — reuse the one already on CodebaseLLMAgent if available,
# otherwise we load constraints ourselves with a simple regex parser.
try:
    from agents.codebase_llm_agent import CodebaseLLMAgent
    LLM_AGENT_AVAILABLE = True
except ImportError:
    LLM_AGENT_AVAILABLE = False
    CodebaseLLMAgent = None

try:
    from agents.codebase_static_agent import StaticAnalyzerAgent
    STATIC_AGENT_AVAILABLE = True
except ImportError:
    STATIC_AGENT_AVAILABLE = False
    StaticAnalyzerAgent = None

try:
    from agents.adapters import (
        ASTComplexityAdapter,
        SecurityAdapter,
        DeadCodeAdapter,
        CallGraphAdapter,
        FunctionMetricsAdapter,
    )
    ADAPTERS_AVAILABLE = True
except ImportError:
    ADAPTERS_AVAILABLE = False

try:
    from utils.common.excel_writer import ExcelWriter, ExcelStyle
    EXCEL_WRITER_AVAILABLE = True
except ImportError:
    EXCEL_WRITER_AVAILABLE = False
    ExcelWriter = None
    ExcelStyle = None

try:
    from utils.common.llm_tools import LLMTools
    LLM_TOOLS_AVAILABLE = True
except ImportError:
    LLM_TOOLS_AVAILABLE = False
    LLMTools = None

try:
    from utils.parsers.global_config_parser import GlobalConfig
    GLOBAL_CONFIG_AVAILABLE = True
except ImportError:
    GLOBAL_CONFIG_AVAILABLE = False
    GlobalConfig = None

try:
    from dependency_builder.config import DependencyBuilderConfig
    DEP_CONFIG_AVAILABLE = True
except ImportError:
    DEP_CONFIG_AVAILABLE = False
    DependencyBuilderConfig = None

# HITL support (optional)
try:
    from hitl import HITLContext, HITL_AVAILABLE
except ImportError:
    HITLContext = None
    HITL_AVAILABLE = False


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class PatchHunk:
    """Represents a single hunk from a unified diff."""

    orig_start: int
    orig_count: int
    new_start: int
    new_count: int
    header: str = ""
    removed_lines: List[str] = field(default_factory=list)
    added_lines: List[str] = field(default_factory=list)
    context_lines: List[str] = field(default_factory=list)
    raw_lines: List[str] = field(default_factory=list)


@dataclass
class PatchFinding:
    """A single issue found in the patched code."""

    file_path: str
    line_number: int
    severity: str
    category: str
    description: str
    title: str = ""
    confidence: str = ""
    suggestion: str = ""
    code_before: str = ""
    code_after: str = ""
    introduced_by_patch: bool = True
    issue_source: str = "patch"
    feedback: str = ""
    constraints: str = ""


# ---------------------------------------------------------------------------
# Patch Agent
# ---------------------------------------------------------------------------

class CodebasePatchAgent:
    """Analyse patches against source files to identify introduced issues.

    Workflow:
    1. Read the original source file.
    2. Parse the unified diff via :meth:`_parse_patch`.
    3. Apply the patch to produce a patched source string.
    4. Run :class:`CodebaseLLMAgent` on **both** original and patched files.
       (Injects Issue Identification Rules from constraints)
    5. Optionally run static analysis adapters on the patched file.
    6. Diff findings to isolate issues **introduced** by the patch.
    7. Write a ``patch_<filename>`` tab to ``detailed_code_review.xlsx``.
    """

    def __init__(
        self,
        file_path: str,
        patch_file: str,
        output_dir: str = "./out",
        config: Optional[Any] = None,
        llm_tools: Optional[Any] = None,
        dep_config: Optional[Any] = None,
        hitl_context: Optional[Any] = None,
        enable_adapters: bool = False,
        verbose: bool = False,
        constraints_dir: str = "agents/constraints",
        exclude_dirs: Optional[List[str]] = None,
        exclude_globs: Optional[List[str]] = None,
        custom_constraints: Optional[List[str]] = None,
        codebase_path: Optional[str] = None,
    ) -> None:
        self.file_path = Path(file_path).resolve()
        self.patch_file = Path(patch_file).resolve()
        self.output_dir = Path(output_dir).resolve()
        self.config = config
        self.llm_tools = llm_tools
        self.dep_config = dep_config
        self.hitl_context = hitl_context
        self.enable_adapters = enable_adapters
        self.verbose = verbose
        self.exclude_dirs = exclude_dirs or []
        self.exclude_globs = exclude_globs or []
        self.custom_constraints = custom_constraints or []

        # Context codebase path — the REAL codebase root used by the inner
        # CodebaseLLMAgent for header resolution, context validation, and
        # call-stack analysis.  Without this, those layers would only see the
        # single file in the temp directory and produce no useful context.
        if codebase_path:
            self.codebase_path = Path(codebase_path).resolve()
        else:
            self.codebase_path = self.file_path.parent

        # Derive useful names
        self.filename = self.file_path.name
        self.filename_stem = self.file_path.stem

        # Constraint Directory Setup
        self.constraints_dir = Path(constraints_dir)
        if not self.constraints_dir.is_absolute():
            # Attempt to resolve relative to CWD first, then fallback to script location
            if not self.constraints_dir.exists():
                # Fallback: check relative to this script's directory
                script_dir = Path(__file__).parent.resolve()
                potential_dir = script_dir / constraints_dir
                if potential_dir.exists():
                    self.constraints_dir = potential_dir

        # Ensure output dir exists
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Setup Logging — use module logger only (do NOT call basicConfig which
        # installs a StreamHandler on the root logger and floods the UI console)
        self.logger = logging.getLogger(__name__)

        # Temp directory for analysis artefacts
        self._temp_dir: Optional[Path] = None

        # --- Context layers (initialised lazily or here) ---
        # ContextValidator: per-chunk false-positive reduction
        self.context_validator = None
        if CTX_VALIDATOR_AVAILABLE:
            try:
                self.context_validator = ContextValidator()
                self.logger.debug("  ContextValidator enabled for patch review")
            except Exception as cv_err:
                self.logger.debug(f"  ContextValidator init failed: {cv_err}")

        # StaticCallStackAnalyzer: cross-function call-chain tracing
        self.call_stack_analyzer = None
        if CALL_STACK_AVAILABLE:
            try:
                self.call_stack_analyzer = StaticCallStackAnalyzer(
                    codebase_root=str(self.codebase_path)
                )
                self.logger.debug("  StaticCallStackAnalyzer enabled for patch review")
            except Exception as csa_err:
                self.logger.debug(f"  StaticCallStackAnalyzer init failed: {csa_err}")

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def run_analysis(
        self,
        excel_path: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Execute the full patch analysis pipeline.

        Args:
            excel_path: Path to ``detailed_code_review.xlsx`` to update.

        Returns:
            Dictionary with analysis results.
        """
        self.logger.info(f"Patch Agent: analysing {self.filename} with {self.patch_file.name}")

        # -- Validate inputs ------------------------------------------------
        if not self.file_path.exists():
            self.logger.error(f"Source file not found: {self.file_path}")
            return {"status": "error", "message": f"Source file not found: {self.file_path}"}

        if not self.patch_file.exists():
            self.logger.error(f"Patch file not found: {self.patch_file}")
            return {"status": "error", "message": f"Patch file not found: {self.patch_file}"}

        try:
            original_content = self.file_path.read_text(encoding="utf-8", errors="replace")
        except Exception as exc:
            self.logger.error(f"Failed to read source file: {exc}")
            return {"status": "error", "message": str(exc)}

        try:
            patch_content = self.patch_file.read_text(encoding="utf-8", errors="replace")
        except Exception as exc:
            self.logger.error(f"Failed to read patch file: {exc}")
            return {"status": "error", "message": str(exc)}

        # -- Strip BOM and normalize line endings ---------------------------
        # BOM prefix (\ufeff) causes ^@@ regex to fail — strip it.
        if patch_content.startswith("\ufeff"):
            patch_content = patch_content[1:]
            self.logger.debug("  Stripped BOM prefix from patch file")
        patch_content = patch_content.replace("\r\n", "\n").replace("\r", "\n")

        # -- Parse and apply patch ------------------------------------------
        hunks = self._parse_patch(patch_content)
        if not hunks:
            self.logger.warning("No hunks found in patch file")
            return {"status": "warning", "message": "No hunks found in patch", "findings": []}

        self.logger.info(f"  Parsed {len(hunks)} hunk(s) from patch")

        try:
            patched_content = self._apply_patch(original_content, hunks)
        except Exception as e:
            self.logger.error(f"Failed to apply patch: {e}")
            return {"status": "error", "message": f"Patch application failed: {e}"}

        # -- Setup temp directories -----------------------------------------
        try:
            self._temp_dir = Path(tempfile.mkdtemp(prefix="cure_patch_"))
            return self._run_pipeline(
                original_content, patched_content, hunks, excel_path
            )
        except Exception as e:
            self.logger.error(f"Pipeline execution failed: {e}")
            return {"status": "error", "message": str(e)}
        finally:
            # Cleanup temp dir safely
            if self._temp_dir and self._temp_dir.exists():
                try:
                    shutil.rmtree(str(self._temp_dir), ignore_errors=True)
                except Exception as e:
                    self.logger.warning(f"Failed to cleanup temp dir: {e}")

    # ------------------------------------------------------------------
    # Pipeline
    # ------------------------------------------------------------------

    def _run_pipeline(
        self,
        original_content: str,
        patched_content: str,
        hunks: List[PatchHunk],
        excel_path: Optional[str],
    ) -> Dict[str, Any]:
        """Run the full analysis pipeline."""

        # Write temp files for analysis
        orig_dir = self._temp_dir / "original"
        patched_dir = self._temp_dir / "patched"
        orig_dir.mkdir(parents=True, exist_ok=True)
        patched_dir.mkdir(parents=True, exist_ok=True)

        orig_file = orig_dir / self.filename
        patched_file = patched_dir / self.filename
        orig_file.write_text(original_content, encoding="utf-8")
        patched_file.write_text(patched_content, encoding="utf-8")

        # -- Compute focus ranges for each version ----------------------------
        # *Exact* hunk ranges (no padding) — used in the prompt to tell the
        # LLM which lines are actually changed by the patch.
        # *Padded* ranges (±20 lines) — used to extract the code sent to the
        # LLM so it has surrounding function-level context.
        _PAD = 20

        # Exact ranges (what the patch actually changes)
        orig_exact = [
            (h.orig_start, h.orig_start + max(h.orig_count - 1, 0))
            for h in hunks
        ]
        patched_exact = [
            (h.new_start, h.new_start + max(h.new_count - 1, 0))
            for h in hunks
        ]

        # Padded ranges (code extraction window)
        orig_focus = [
            (max(1, h.orig_start - _PAD), h.orig_start + h.orig_count + _PAD)
            for h in hunks
        ]
        patched_focus = [
            (max(1, h.new_start - _PAD), h.new_start + h.new_count + _PAD)
            for h in hunks
        ]

        # -- Run LLM analysis on both versions ------------------------------
        original_issues: List[Dict] = []
        patched_issues: List[Dict] = []

        if self.llm_tools and PATCH_PROMPT_AVAILABLE:
            self.logger.info("  Running patch LLM analysis on original file...")
            original_issues = self._run_patch_llm_analysis(
                original_content, self.filename, "original",
                focus_line_ranges=orig_focus,
                exact_hunk_ranges=orig_exact,
            )
            self.logger.info(f"  Original: {len(original_issues)} issue(s) found")

            self.logger.info("  Running patch LLM analysis on patched file...")
            patched_issues = self._run_patch_llm_analysis(
                patched_content, self.filename, "patched",
                focus_line_ranges=patched_focus,
                exact_hunk_ranges=patched_exact,
            )
            self.logger.info(f"  Patched: {len(patched_issues)} issue(s) found")
        else:
            reason = []
            if not self.llm_tools:
                reason.append("LLMTools not available")
            if not PATCH_PROMPT_AVAILABLE:
                reason.append("PATCH_REVIEW_PROMPT not found")
            self.logger.warning(f"  Skipping LLM analysis — {', '.join(reason)}")

        # -- Run static adapters on patched file (optional — user must enable) --
        static_issues: List[Dict] = []
        adapter_raw_results: Dict[str, List[Dict]] = {}
        if self.enable_adapters and ADAPTERS_AVAILABLE:
            self.logger.info("  Running static adapters on patched file...")
            static_issues, adapter_raw_results = self._run_static_analysis(
                str(patched_dir), self.filename,
                focus_line_ranges=patched_exact,  # use EXACT hunk ranges
            )
            self.logger.info(f"  Static analysis: {len(static_issues)} issue(s) found")

        # Merge patched_issues + static_issues
        all_patched_issues = patched_issues + static_issues

        # -- Diff findings --------------------------------------------------
        new_findings = self._diff_findings(original_issues, all_patched_issues, hunks)
        self.logger.info(f"  New issues introduced by patch: {len(new_findings)}")

        # -- Write to Excel -------------------------------------------------
        final_excel = excel_path or str(self.output_dir / "detailed_code_review.xlsx")
        self._update_excel(final_excel, new_findings, adapter_raw_results)

        return {
            "status": "success",
            "filename": self.filename,
            "patch_file": str(self.patch_file),
            "original_issue_count": len(original_issues),
            "patched_issue_count": len(all_patched_issues),
            "new_issue_count": len(new_findings),
            "findings": [self._finding_to_dict(f) for f in new_findings],
            "excel_path": final_excel,
            "hunks_parsed": len(hunks),
        }

    # ------------------------------------------------------------------
    # Patch parsing — supports unified, context, normal, and combined
    # ------------------------------------------------------------------

    # Unified diff:   @@ -start,count +start,count @@
    _UNIFIED_HUNK_RE = re.compile(
        r"^@@\s+-(\d+)(?:,(\d+))?\s+\+(\d+)(?:,(\d+))?\s+@@(.*)$"
    )
    # Combined diff:  @@@ -start,count -start,count +start,count @@@
    _COMBINED_HUNK_RE = re.compile(
        r"^@@@\s+.*?\+(\d+)(?:,(\d+))?\s+@@@(.*)$"
    )
    # Context diff section separator
    _CTX_SEP_RE = re.compile(r"^\*{15,}")
    # Context diff original range: *** start,end ****
    _CTX_ORIG_RE = re.compile(r"^\*\*\*\s+(\d+)(?:,(\d+))?\s+\*{4}")
    # Context diff new range:      --- start,end ----
    _CTX_NEW_RE = re.compile(r"^---\s+(\d+)(?:,(\d+))?\s+-{4}")
    # Normal diff commands: NUMaNUM, NUMcNUM, NUMdNUM (with optional ranges)
    _NORMAL_CMD_RE = re.compile(
        r"^(\d+)(?:,(\d+))?([acd])(\d+)(?:,(\d+))?$"
    )

    def _detect_diff_format(self, patch_text: str) -> str:
        """Auto-detect the diff format from the patch text.

        Returns one of: ``"unified"``, ``"context"``, ``"normal"``,
        ``"combined"``, or ``"unknown"``.
        """
        for line in patch_text.splitlines()[:100]:
            if self._COMBINED_HUNK_RE.match(line):
                return "combined"
            if self._UNIFIED_HUNK_RE.match(line):
                return "unified"
            if self._CTX_SEP_RE.match(line):
                return "context"
            if self._NORMAL_CMD_RE.match(line):
                return "normal"
        return "unknown"

    def _parse_patch(self, patch_text: str) -> List[PatchHunk]:
        """Auto-detect format and parse into :class:`PatchHunk` objects.

        Supported formats:
          - **Unified diff**  (``diff -u``, ``git diff``) — ``@@`` markers
          - **Context diff**  (``diff -c``) — ``***`` / ``---`` markers
          - **Normal diff**   (``diff``) — ``NUMaNUM`` / ``NUMcNUM`` / ``NUMdNUM``
          - **Combined diff** (``git diff`` merge conflicts) — ``@@@`` markers
        """
        fmt = self._detect_diff_format(patch_text)
        self.logger.info(f"  Detected diff format: {fmt}")

        if fmt == "unified":
            return self._parse_unified(patch_text)
        if fmt == "context":
            return self._parse_context(patch_text)
        if fmt == "normal":
            return self._parse_normal(patch_text)
        if fmt == "combined":
            return self._parse_combined(patch_text)

        # Fallback: try unified anyway (most common)
        self.logger.warning("  Unknown diff format — falling back to unified parser")
        return self._parse_unified(patch_text)

    # ── Unified diff parser ──────────────────────────────────────────

    def _parse_unified(self, patch_text: str) -> List[PatchHunk]:
        """Parse a unified diff (``diff -u`` / ``git diff``)."""
        hunks: List[PatchHunk] = []
        current_hunk: Optional[PatchHunk] = None

        for line in patch_text.splitlines():
            m = self._UNIFIED_HUNK_RE.match(line)
            if m:
                if current_hunk is not None:
                    hunks.append(current_hunk)
                current_hunk = PatchHunk(
                    orig_start=int(m.group(1)),
                    orig_count=int(m.group(2) or 1),
                    new_start=int(m.group(3)),
                    new_count=int(m.group(4) or 1),
                    header=m.group(5).strip(),
                )
                continue

            if current_hunk is None:
                continue

            current_hunk.raw_lines.append(line)
            if line.startswith("-"):
                current_hunk.removed_lines.append(line[1:])
            elif line.startswith("+"):
                current_hunk.added_lines.append(line[1:])
            elif line.startswith(" ") or line == "":
                current_hunk.context_lines.append(
                    line[1:] if line.startswith(" ") else line
                )

        if current_hunk is not None:
            hunks.append(current_hunk)
        return hunks

    # ── Context diff parser ──────────────────────────────────────────

    def _parse_context(self, patch_text: str) -> List[PatchHunk]:
        """Parse a context diff (``diff -c``).

        Context diff structure::

            *** file1.c  timestamp
            --- file2.c  timestamp
            ***************
            *** 10,15 ****
              context line
            ! changed line (original)
            - removed line
              context line
            --- 10,16 ----
              context line
            ! changed line (new)
            + added line
              context line
        """
        hunks: List[PatchHunk] = []
        lines = patch_text.splitlines()
        i = 0
        n = len(lines)

        while i < n:
            # Look for *************** separator
            if not self._CTX_SEP_RE.match(lines[i]):
                i += 1
                continue

            i += 1
            if i >= n:
                break

            # --- Parse original section: *** start,end ****
            m_orig = self._CTX_ORIG_RE.match(lines[i])
            if not m_orig:
                continue
            orig_start = int(m_orig.group(1))
            orig_end = int(m_orig.group(2) or orig_start)
            orig_count = orig_end - orig_start + 1
            i += 1

            # Collect original section lines
            removed: List[str] = []
            orig_context: List[str] = []
            orig_raw: List[str] = []
            while i < n:
                ln = lines[i]
                if self._CTX_NEW_RE.match(ln):
                    break
                if self._CTX_SEP_RE.match(ln):
                    break
                orig_raw.append(ln)
                if ln.startswith("- ") or ln.startswith("-\t"):
                    removed.append(ln[2:])
                elif ln.startswith("! "):
                    removed.append(ln[2:])
                elif ln.startswith("  ") or ln.startswith("\t"):
                    orig_context.append(ln[2:] if ln.startswith("  ") else ln[1:])
                i += 1

            # --- Parse new section: --- start,end ----
            if i >= n:
                break
            m_new = self._CTX_NEW_RE.match(lines[i])
            if not m_new:
                # No new section — treat as pure deletion
                hunks.append(PatchHunk(
                    orig_start=orig_start,
                    orig_count=orig_count,
                    new_start=orig_start,
                    new_count=0,
                    header="",
                    removed_lines=removed,
                    added_lines=[],
                    context_lines=orig_context,
                    raw_lines=orig_raw,
                ))
                continue

            new_start = int(m_new.group(1))
            new_end = int(m_new.group(2) or new_start)
            new_count = new_end - new_start + 1
            i += 1

            # Collect new section lines
            added: List[str] = []
            new_context: List[str] = []
            raw_lines: List[str] = list(orig_raw)
            while i < n:
                ln = lines[i]
                if self._CTX_SEP_RE.match(ln):
                    break
                if self._CTX_ORIG_RE.match(ln):
                    break
                raw_lines.append(ln)
                if ln.startswith("+ ") or ln.startswith("+\t"):
                    added.append(ln[2:])
                elif ln.startswith("! "):
                    added.append(ln[2:])
                elif ln.startswith("  ") or ln.startswith("\t"):
                    new_context.append(ln[2:] if ln.startswith("  ") else ln[1:])
                i += 1

            # Build unified-style raw_lines for _apply_patch compatibility
            unified_raw: List[str] = []
            for r in removed:
                unified_raw.append(f"-{r}")
            for c in orig_context:
                unified_raw.append(f" {c}")
            for a in added:
                unified_raw.append(f"+{a}")

            hunks.append(PatchHunk(
                orig_start=orig_start,
                orig_count=orig_count,
                new_start=new_start,
                new_count=new_count,
                header="",
                removed_lines=removed,
                added_lines=added,
                context_lines=list(set(orig_context + new_context)),
                raw_lines=unified_raw,
            ))

        return hunks

    # ── Normal diff parser ───────────────────────────────────────────

    def _parse_normal(self, patch_text: str) -> List[PatchHunk]:
        """Parse a normal diff (plain ``diff`` output).

        Normal diff format::

            5a6,7        ← add after line 5: lines 6-7 in new file
            > added line1
            > added line2
            10,12c10,11  ← change lines 10-12 to lines 10-11
            < old line1
            < old line2
            < old line3
            ---
            > new line1
            > new line2
            15d14        ← delete line 15 (was before line 14 in new)
            < deleted line
        """
        hunks: List[PatchHunk] = []
        lines = patch_text.splitlines()
        i = 0
        n = len(lines)

        while i < n:
            m = self._NORMAL_CMD_RE.match(lines[i])
            if not m:
                i += 1
                continue

            orig_s = int(m.group(1))
            orig_e = int(m.group(2) or orig_s)
            cmd = m.group(3)          # 'a', 'c', or 'd'
            new_s = int(m.group(4))
            new_e = int(m.group(5) or new_s)
            i += 1

            removed: List[str] = []
            added: List[str] = []
            raw_lines: List[str] = []

            # For 'a' (add): orig side has 0 lines removed
            # For 'd' (delete): new side has 0 lines added
            # For 'c' (change): both sides present, separated by '---'

            if cmd == "a":
                # Collect '>' lines (additions)
                while i < n and lines[i].startswith("> "):
                    content = lines[i][2:]
                    added.append(content)
                    raw_lines.append(f"+{content}")
                    i += 1
                hunks.append(PatchHunk(
                    orig_start=orig_s + 1,  # 'a' means "after orig_s"
                    orig_count=0,
                    new_start=new_s,
                    new_count=new_e - new_s + 1,
                    header="",
                    removed_lines=[],
                    added_lines=added,
                    context_lines=[],
                    raw_lines=raw_lines,
                ))

            elif cmd == "d":
                # Collect '<' lines (deletions)
                while i < n and lines[i].startswith("< "):
                    content = lines[i][2:]
                    removed.append(content)
                    raw_lines.append(f"-{content}")
                    i += 1
                hunks.append(PatchHunk(
                    orig_start=orig_s,
                    orig_count=orig_e - orig_s + 1,
                    new_start=new_s,
                    new_count=0,
                    header="",
                    removed_lines=removed,
                    added_lines=[],
                    context_lines=[],
                    raw_lines=raw_lines,
                ))

            elif cmd == "c":
                # Collect '<' lines (original / removed)
                while i < n and lines[i].startswith("< "):
                    content = lines[i][2:]
                    removed.append(content)
                    raw_lines.append(f"-{content}")
                    i += 1

                # Skip '---' separator
                if i < n and lines[i] == "---":
                    i += 1

                # Collect '>' lines (new / added)
                while i < n and lines[i].startswith("> "):
                    content = lines[i][2:]
                    added.append(content)
                    raw_lines.append(f"+{content}")
                    i += 1

                hunks.append(PatchHunk(
                    orig_start=orig_s,
                    orig_count=orig_e - orig_s + 1,
                    new_start=new_s,
                    new_count=new_e - new_s + 1,
                    header="",
                    removed_lines=removed,
                    added_lines=added,
                    context_lines=[],
                    raw_lines=raw_lines,
                ))

        return hunks

    # ── Combined diff parser ─────────────────────────────────────────

    def _parse_combined(self, patch_text: str) -> List[PatchHunk]:
        """Parse a combined diff (``git diff`` on merge conflicts).

        Combined diffs use ``@@@`` markers and have two original columns.
        We treat them as a single unified diff using only the final
        (merged) column.
        """
        hunks: List[PatchHunk] = []
        current_hunk: Optional[PatchHunk] = None

        for line in patch_text.splitlines():
            m = self._COMBINED_HUNK_RE.match(line)
            if m:
                if current_hunk is not None:
                    hunks.append(current_hunk)
                new_start = int(m.group(1))
                new_count = int(m.group(2) or 1)
                current_hunk = PatchHunk(
                    orig_start=new_start,
                    orig_count=new_count,
                    new_start=new_start,
                    new_count=new_count,
                    header=m.group(3).strip(),
                )
                continue

            if current_hunk is None:
                continue

            current_hunk.raw_lines.append(line)

            # Combined diffs prefix lines with two columns (e.g., '+ ', ' +', '++', '  ')
            # We simplify: lines starting with '++' or '+ ' are additions,
            # lines starting with '--' are removals, everything else is context.
            if line.startswith("++") or line.startswith("+ "):
                current_hunk.added_lines.append(line[2:] if len(line) > 2 else "")
            elif line.startswith("--"):
                current_hunk.removed_lines.append(line[2:] if len(line) > 2 else "")
            else:
                ctx = line[2:] if len(line) > 2 else (line[1:] if len(line) > 1 else "")
                current_hunk.context_lines.append(ctx)

        if current_hunk is not None:
            hunks.append(current_hunk)
        return hunks

    # ------------------------------------------------------------------
    # Patch application
    # ------------------------------------------------------------------

    def _apply_patch(self, source: str, hunks: List[PatchHunk]) -> str:
        """Apply parsed hunks to the original source to reconstruct patched content.

        Uses a line-based approach: replaces each hunk's original region
        with the new region in order, adjusting offsets as we go.
        """
        lines = source.splitlines(keepends=True)
        offset = 0  # cumulative line offset from previous hunk applications

        for hunk in hunks:
            # Convert to 0-based indexing
            start = max(0, hunk.orig_start - 1 + offset)
            end = start + hunk.orig_count

            # Build replacement lines from the raw diff
            new_lines: List[str] = []
            for raw_line in hunk.raw_lines:
                if raw_line.startswith("+"):
                    new_lines.append(raw_line[1:] + "\n")
                elif raw_line.startswith(" ") or raw_line == "":
                    # Handle preserved lines
                    content = raw_line[1:] if raw_line.startswith(" ") else raw_line
                    new_lines.append(content + "\n")
                # Lines starting with "-" are removed (not added to new_lines)

            # Safety check: if start is beyond EOF, append (though likely a bad patch)
            if start > len(lines):
                 lines.extend(new_lines)
            else:
                 # Replace the region
                 lines[start:end] = new_lines

            # Update offset for next hunk
            offset += len(new_lines) - hunk.orig_count

        return "".join(lines)

    # ------------------------------------------------------------------
    # LLM analysis (self-contained — no CodebaseLLMAgent dependency)
    # ------------------------------------------------------------------

    # Padding (lines) around each hunk for function-level context
    _HUNK_PAD = 20

    def _run_patch_llm_analysis(
        self,
        file_content: str,
        filename: str,
        label: str,
        focus_line_ranges: Optional[List[tuple]] = None,
        exact_hunk_ranges: Optional[List[tuple]] = None,
    ) -> List[Dict]:
        """Self-contained LLM analysis scoped to patch hunk regions.

        Instead of delegating to :class:`CodebaseLLMAgent` (which scans
        the entire file), this method:

        1. Extracts only the code surrounding each hunk from *file_content*.
        2. Gathers 4-layer context from the **real codebase** (header
           context, context validation, call-stack analysis, constraints).
        3. Builds a prompt using ``PATCH_REVIEW_PROMPT``.
        4. Calls ``self.llm_tools.llm_call()`` directly.
        5. Parses the ``---ISSUE---`` response blocks.

        Args:
            focus_line_ranges: Padded ranges (±20 lines) used for code extraction.
            exact_hunk_ranges: Exact hunk line ranges (no padding) — the lines
                actually changed by the patch. Used in the prompt to tell the
                LLM precisely which lines to review.

        Returns a list of issue dicts.
        """
        if not self.llm_tools:
            return []
        if not PATCH_PROMPT_AVAILABLE:
            self.logger.warning("PATCH_REVIEW_PROMPT not available — skipping")
            return []

        all_issues: List[Dict] = []
        all_lines = file_content.splitlines()
        total_lines = len(all_lines)

        # --- Build file-level context layers (resolved once per file) ------
        # 1. Constraints (Issue Identification Rules from .md files)
        constraints_context = self._load_constraints_for_file(filename)

        # 2. Header includes (resolved once, context built per-chunk below)
        file_includes = self._resolve_file_includes(filename)

        # 3. External dependencies via CCLS (if available and indexed)
        #    — populated per-chunk below via _fetch_dependency_context()

        # --- Iterate over focus ranges ---
        ranges = focus_line_ranges or [(1, total_lines)]
        for rng_idx, (rng_start, rng_end) in enumerate(ranges):
            # Clamp to file bounds
            chunk_start = max(1, rng_start)
            chunk_end = min(total_lines, rng_end)
            if chunk_start >= chunk_end:
                continue

            # Extract chunk (1-based inclusive)
            chunk_lines = all_lines[chunk_start - 1 : chunk_end]
            chunk_text = "\n".join(chunk_lines)
            chunk_line_count = len(chunk_lines)

            # Numbered code block with visual markers for changed vs context lines.
            # Lines within exact hunk ranges get a '>>>' prefix so the LLM
            # can clearly distinguish changed lines from surrounding context.
            def _in_exact_range(line_num: int) -> bool:
                """Check if line_num falls within any exact hunk range."""
                if not exact_hunk_ranges:
                    return True  # no exact ranges = treat all as changed
                for ex_start, ex_end in exact_hunk_ranges:
                    if ex_start <= line_num <= ex_end:
                        return True
                return False

            numbered = []
            for idx, line in enumerate(chunk_lines):
                abs_line = chunk_start + idx
                marker = ">>>" if _in_exact_range(abs_line) else "   "
                numbered.append(f"{marker} {abs_line:5d} | {line}")
            numbered_code_block = "\n".join(numbered)

            # ==============================================================
            # Per-chunk context layers (matches CodebaseLLMAgent exactly)
            # ==============================================================

            # Layer 2a: External Dependencies via CCLS
            dependency_context = self._fetch_dependency_context(
                filename, chunk_start, chunk_start + chunk_line_count
            )

            # Layer 2b: Header Context (struct/enum/macro definitions
            #           from included headers, filtered to this chunk)
            header_context = ""
            if HEADER_CTX_AVAILABLE and hasattr(self, '_header_builder') and file_includes:
                try:
                    header_context = self._header_builder.build_context_for_chunk(
                        chunk_text, file_includes
                    ) or ""
                    if header_context:
                        self.logger.debug(
                            f"    Header context for chunk {rng_idx+1}: "
                            f"{len(header_context)} chars injected"
                        )
                except Exception as hctx_err:
                    self.logger.debug(f"    Header context build failed: {hctx_err}")

            # Layer 2c: Context Validation (per-chunk false positive reduction)
            validation_context = ""
            if CTX_VALIDATOR_AVAILABLE and self.context_validator:
                try:
                    val_report = self.context_validator.analyze_chunk(
                        chunk_text, str(self.file_path), file_content, chunk_start
                    )
                    validation_context = val_report.format_summary(max_chars=10000)
                    if validation_context:
                        self.logger.debug(
                            f"    Validation context for chunk {rng_idx+1}: "
                            f"{len(validation_context)} chars, "
                            f"{len(val_report.validations)} symbols checked"
                        )
                except Exception as cv_err:
                    self.logger.debug(f"  Context validation failed: {cv_err}")

            # Layer 2d: Call Stack Context (cross-function call chain tracing)
            chunk_call_stack = ""
            if CALL_STACK_AVAILABLE and self.call_stack_analyzer:
                try:
                    chunk_call_stack = self.call_stack_analyzer.analyze_chunk(
                        chunk_text, str(self.file_path), file_content, chunk_start
                    )
                    if chunk_call_stack:
                        self.logger.debug(
                            f"    Call stack context for chunk {rng_idx+1}: "
                            f"{len(chunk_call_stack)} chars injected"
                        )
                except Exception as csa_err:
                    self.logger.debug(f"  Call stack analysis failed: {csa_err}")

            # ==============================================================
            # Assemble context header (same ordering as CodebaseLLMAgent)
            # ==============================================================
            context_header = ""

            if dependency_context:
                context_header += (
                    f"\n// ... [CONTEXT: External Dependencies & Definitions] ...\n"
                    f"{dependency_context}\n"
                    f"// ... [End External Context] ...\n"
                )

            if header_context:
                context_header += f"\n{header_context}\n"

            if validation_context:
                context_header += f"\n{validation_context}\n"

            if chunk_call_stack:
                context_header += f"\n{chunk_call_stack}\n"

            # --- Constraints ---
            prompt_constraints = ""
            if constraints_context:
                prompt_constraints = f"""
                ========================================
                MANDATORY IDENTIFICATION RULES (IGNORE FALSE POSITIVES)
                ========================================
                {constraints_context}
                ========================================
                """

            # --- Patch line ranges block ---
            # Use EXACT hunk ranges (no padding) so the LLM knows precisely
            # which lines were changed.  The padded focus_line_ranges are only
            # used for code extraction — the LLM should NOT flag issues on
            # the buffer/context lines.
            exact_ranges_for_prompt = exact_hunk_ranges or focus_line_ranges or ranges
            range_strs = ", ".join(
                f"{s}-{e}" for s, e in exact_ranges_for_prompt
            )
            patch_scope_block = f"""
            ========================================
            PATCH LINE RANGES (ONLY flag issues on these exact lines):
              [{range_strs}]
            Lines marked with '>>>' in the code below are CHANGED lines.
            Lines marked with '   ' are CONTEXT ONLY — do NOT flag issues on them.
            ========================================
            """

            # --- Final chunk text with context prepended ---
            final_chunk_text = (
                f"{context_header}\n"
                f"// ... [PATCH CHUNK: {filename} Lines {chunk_start}-{chunk_end}] ...\n"
                f"{numbered_code_block}"
            )

            # --- Build prompt ---
            final_prompt = f"""
            {PATCH_REVIEW_PROMPT}

            {prompt_constraints}

            {patch_scope_block}

            TARGET SOURCE CODE ({filename} - Hunk region {rng_idx+1}/{len(ranges)}):
            ```cpp
            {final_chunk_text}
            ```
            """

            # --- HITL: augment prompt with feedback context ---
            if self.hitl_context:
                try:
                    final_prompt = self.hitl_context.augment_prompt(
                        original_prompt=final_prompt,
                        issue_type="code_quality",
                        file_path=filename,
                        agent_type="patch_agent",
                    )
                except Exception as hitl_err:
                    self.logger.debug(f"  HITL augment failed: {hitl_err}")

            # --- Debug: dump prompt to {output_dir}/prompt_dumps/ ---
            #     Always dump for patch reviews (not just DEBUG level)
            #     so users can inspect what was sent to the LLM.
            try:
                dump_dir = os.path.join(str(self.output_dir), "prompt_dumps")
                os.makedirs(dump_dir, exist_ok=True)
                safe_name = filename.replace("/", "__").replace("\\", "__")
                dump_path = os.path.join(
                    dump_dir,
                    f"patch_{safe_name}_{label}_chunk{rng_idx + 1}.txt",
                )
                with open(dump_path, "w", encoding="utf-8") as df:
                    df.write(final_prompt)
                self.logger.info(f"    Prompt dump: {dump_path}")
            except Exception:
                pass  # never fail on debug dump

            # --- LLM call ---
            try:
                response = self.llm_tools.llm_call(final_prompt)
            except Exception as llm_err:
                self.logger.warning(f"LLM call failed for {label} chunk {rng_idx+1}: {llm_err}")
                continue

            # --- Parse response ---
            parsed = self._parse_patch_llm_response(
                response, filename, chunk_text, chunk_start, chunk_line_count
            )
            all_issues.extend(parsed)

        return all_issues

    # ------------------------------------------------------------------
    # Context helpers
    # ------------------------------------------------------------------

    def _resolve_file_includes(self, filename: str) -> list:
        """Resolve #include directives once per file.

        Returns a list of include objects that can be passed to
        ``HeaderContextBuilder.build_context_for_chunk()`` per chunk.
        Mirrors the once-per-file resolution in :class:`CodebaseLLMAgent`.
        """
        if not HEADER_CTX_AVAILABLE:
            return []
        try:
            if not hasattr(self, '_header_builder'):
                config_dict = {}
                if self.config:
                    try:
                        config_dict = {
                            "include_paths": self.config.get("context.include_paths") or [],
                            "max_header_depth": self.config.get("context.max_header_depth") or 2,
                            "max_context_chars": self.config.get("context.max_context_chars") or 6000,
                            "exclude_system_headers": self.config.get("context.exclude_system_headers", True),
                        }
                    except Exception:
                        pass
                self._header_builder = HeaderContextBuilder(
                    codebase_root=str(self.codebase_path),
                    **config_dict,
                )
            includes = self._header_builder.resolve_includes(str(self.file_path))
            if includes:
                resolved = [inc for inc in includes if inc.resolved]
                self.logger.debug(
                    f"    Header includes for {filename}: "
                    f"{len(includes)} total, {len(resolved)} resolved"
                )
                return includes
            else:
                self.logger.debug(f"    No #include directives found in {filename}")
        except Exception as hdr_err:
            self.logger.debug(f"  Header include resolution failed: {hdr_err}")
        return []

    def _fetch_dependency_context(
        self, rel_path: str, start_line: int, end_line: int
    ) -> str:
        """Fetch CCLS external dependency context for a chunk, if available.

        Mirrors :meth:`CodebaseLLMAgent._fetch_chunk_dependencies`.
        Returns an empty string if CCLS is not configured.
        """
        if not self.dep_config:
            return ""
        try:
            from dependency_builder.dependency_service import DependencyService
            if not hasattr(self, '_dep_service'):
                self._dep_service = DependencyService(config=self.dep_config)
            response = self._dep_service.perform_fetch(
                project_root=str(self.codebase_path),
                output_dir=str(self.output_dir),
                codebase_identifier=self.codebase_path.name,
                endpoint_type="fetch_dependencies_by_file",
                file_name=rel_path,
                start=start_line,
                end=end_line,
                level=1,
            )
            data = response.get("data", [])
            if not data:
                return ""
            context_parts = []
            if isinstance(data, list):
                for item in data[:10]:
                    if isinstance(item, dict):
                        name = item.get("name", "Unknown")
                        snippet = (item.get("definition") or item.get("snippet") or "").strip()
                        if snippet:
                            context_parts.append(f"// DEP: {name}\n{snippet}")
            return "\n\n".join(context_parts)
        except ImportError:
            self.logger.debug("  CCLS DependencyService not available")
        except Exception as dep_err:
            self.logger.debug(f"  Dependency context fetch failed: {dep_err}")
        return ""

    def _load_constraints_for_file(self, filename: str) -> str:
        """Load Issue Identification Rules from constraint files."""
        if not self.constraints_dir.exists():
            return ""
        try:
            # Reuse CodebaseLLMAgent's constraint loader if available
            if LLM_AGENT_AVAILABLE and hasattr(CodebaseLLMAgent, '_load_constraints'):
                # Build a lightweight instance just for constraint loading
                # This is safe — _load_constraints is a pure file-reader
                dummy = object.__new__(CodebaseLLMAgent)
                dummy.constraints_dir = self.constraints_dir
                dummy.custom_constraints = self.custom_constraints
                dummy.logger = self.logger
                return dummy._load_constraints(filename, section_keyword="Issue Identification Rules")
            else:
                # Fallback: simple regex loader
                return self._simple_load_constraints(filename)
        except Exception as exc:
            self.logger.debug(f"  Constraints load failed: {exc}")
            return ""

    def _simple_load_constraints(self, filename: str) -> str:
        """Fallback: load constraints using a simple regex parser."""
        parts = []
        if not self.constraints_dir.exists():
            return ""

        md_files = sorted(self.constraints_dir.glob("*.md"))
        for custom in (self.custom_constraints or []):
            p = Path(custom)
            if p.exists() and p not in md_files:
                md_files.append(p)

        section_re = re.compile(
            r"^##\s+Issue Identification Rules",
            re.IGNORECASE | re.MULTILINE,
        )
        for md_path in md_files:
            try:
                text = md_path.read_text(encoding="utf-8", errors="replace")
                m = section_re.search(text)
                if m:
                    # Extract from section header to next ## or end
                    rest = text[m.end():]
                    next_section = re.search(r"^##\s+", rest, re.MULTILINE)
                    if next_section:
                        section_text = rest[:next_section.start()].strip()
                    else:
                        section_text = rest.strip()
                    if section_text:
                        parts.append(f"// Constraints from {md_path.name}:\n{section_text}")
            except Exception:
                continue
        return "\n\n".join(parts)

    # ------------------------------------------------------------------
    # LLM response parser (self-contained, matches ---ISSUE--- format)
    # ------------------------------------------------------------------

    def _parse_patch_llm_response(
        self, response: str, file_path: str,
        chunk_text: str, start_line: int, chunk_line_count: int,
    ) -> List[Dict]:
        """Parse ``---ISSUE---`` blocks from the LLM response.

        Returns a list of dicts with keys matching the patch agent's
        expected format: file_path, line_number, severity, category,
        description, code, fixed_code, source.
        """
        issues: List[Dict] = []
        raw_blocks = response.split("---ISSUE---")

        for block in raw_blocks:
            block = block.strip()
            if not block or "No issues found" in block:
                continue

            patterns = {
                "Title": r"Title:\s*(.+)",
                "Severity": r"Severity:\s*(.+)",
                "Confidence": r"Confidence:\s*(.+)",
                "Category": r"Category:\s*(.+)",
                "Description": r"Description:\s*(.+)",
                "Suggestion": r"Suggestion:\s*(.+)",
            }

            data: Dict[str, str] = {}
            for key, pat in patterns.items():
                m = re.search(pat, block, re.IGNORECASE)
                data[key] = m.group(1).strip() if m else "N/A"

            if data.get("Title") == "N/A":
                continue

            # Code snippet
            code_match = re.search(r"Code:\s*```(?:\w+)?\n(.*?)\n```", block, re.DOTALL)
            if not code_match:
                code_match = re.search(r"Code:\s*(.+?)(?=\nFixed_Code:|$)", block, re.DOTALL)
            raw_code = code_match.group(1).strip() if code_match else ""

            # Fixed code
            fixed_match = re.search(r"Fixed_Code:\s*```(?:\w+)?\n(.*?)\n```", block, re.DOTALL)
            if not fixed_match:
                fixed_match = re.search(r"Fixed_Code:\s*(.+?)(?=$)", block, re.DOTALL)
            fixed_code = fixed_match.group(1).strip() if fixed_match else ""

            # --- Anchor logic: resolve line number ---
            calculated_line = 0
            found_by_anchor = False
            if raw_code:
                idx = chunk_text.find(raw_code)
                if idx == -1:
                    first_line = raw_code.split('\n')[0].strip()
                    if len(first_line) > 10:
                        idx = chunk_text.find(first_line)
                if idx != -1:
                    newlines_before = chunk_text[:idx].count('\n')
                    calculated_line = start_line + newlines_before
                    found_by_anchor = True

            if not found_by_anchor:
                line_match = re.search(r"Line\D*(\d+)", block, re.IGNORECASE)
                if line_match:
                    raw_val = int(line_match.group(1))
                    if raw_val < chunk_line_count and raw_val < start_line:
                        calculated_line = start_line + raw_val - 1
                    else:
                        calculated_line = raw_val
                else:
                    calculated_line = start_line

            issues.append({
                "file_path": file_path,
                "line_number": calculated_line,
                "title": data.get("Title", ""),
                "severity": data.get("Severity", "medium").lower(),
                "confidence": data.get("Confidence", ""),
                "category": data.get("Category", ""),
                "description": data.get("Description", ""),
                "suggestion": data.get("Suggestion", ""),
                "code": raw_code,
                "fixed_code": fixed_code,
                "source": f"patch_llm",
            })

        return issues

    # ------------------------------------------------------------------
    # Static analysis
    # ------------------------------------------------------------------

    def _run_static_analysis(
        self, temp_dir: str, filename: str,
        focus_line_ranges: Optional[List[tuple]] = None,
    ) -> Tuple[List[Dict], Dict[str, List[Dict]]]:
        """Run static analysis adapters on the patched file.

        When ``focus_line_ranges`` is provided, only findings whose line
        number falls within one of the (start, end) ranges are kept.
        This prevents the adapters (which always scan the whole file)
        from flooding the output with pre-existing issues.

        Returns:
            Tuple of (merged_issues, per_adapter_results) where
            per_adapter_results maps adapter name → list of raw detail dicts
            for separate Excel tabs.
        """
        if not ADAPTERS_AVAILABLE:
            return [], {}

        def _in_focus(line: int) -> bool:
            """Return True if *line* falls within any exact hunk range."""
            if not focus_line_ranges:
                return True  # no filter → keep everything
            for fstart, fend in focus_line_ranges:
                if fstart <= line <= fend:
                    return True
            return False

        issues: List[Dict] = []
        adapter_raw: Dict[str, List[Dict]] = {}
        try:
            # Build a minimal file cache for adapters
            from agents.core.file_processor import FileProcessor

            processor = FileProcessor(
                codebase_path=temp_dir,
                exclude_dirs=[],
            )
            file_cache = processor.process_files()

            adapters = [
                ("ast_complexity", ASTComplexityAdapter()),
                ("security", SecurityAdapter()),
            ]

            for name, adapter in adapters:
                try:
                    result = adapter.analyze(
                        file_cache, ccls_navigator=None, dependency_graph={}
                    )
                    if result.get("tool_available", False):
                        # Extract issues from adapter results — use 'details'
                        # (per-finding dicts from BaseStaticAdapter._make_detail)
                        # falling back to 'findings' / 'issues' for compat.
                        raw_findings = (
                            result.get("details")
                            or result.get("findings")
                            or result.get("issues", [])
                        )
                        filtered_details: List[Dict] = []
                        for finding in raw_findings:
                            if not isinstance(finding, dict):
                                continue
                            line_num = finding.get("line", 0)
                            # --- Patch scope filter (exact hunk ranges) ---
                            if not _in_focus(line_num):
                                continue
                            filtered_details.append(finding)
                            issues.append({
                                "file_path": finding.get("file", filename),
                                "line_number": line_num,
                                "title": finding.get("description", finding.get("message", ""))[:80],
                                "severity": finding.get("severity", "medium"),
                                "confidence": "CERTAIN",
                                "category": f"static_{name}",
                                "description": finding.get("description", finding.get("message", "")),
                                "suggestion": "",
                                "code": finding.get("code", ""),
                                "fixed_code": "",
                                "source": f"static_{name}",
                            })
                        if filtered_details:
                            adapter_raw[name] = filtered_details
                except Exception as exc:
                    self.logger.warning(f"Adapter {name} failed: {exc}")

        except ImportError:
            self.logger.warning("FileProcessor not available — skipping static analysis")
        except Exception as exc:
            self.logger.warning(f"Static analysis failed: {exc}")

        self.logger.info(
            f"  Static adapters: {len(issues)} issue(s) after patch-scope filtering"
        )
        return issues, adapter_raw

    # ------------------------------------------------------------------
    # Findings diff
    # ------------------------------------------------------------------

    @staticmethod
    def _fingerprint_issue(issue: Dict) -> str:
        """Create a fingerprint for an issue to enable deduplication.

        Uses: (filename, line_range_bucket, category, description_prefix).
        Line numbers are bucketed into ranges of 5 to handle minor drift.
        """
        filename = Path(issue.get("file_path", "")).name
        line = issue.get("line_number", 0)
        line_bucket = (line // 5) * 5  # bucket into groups of 5
        category = issue.get("category", "").lower().strip()
        desc = issue.get("description", "")[:80].lower().strip()

        return f"{filename}|{line_bucket}|{category}|{desc}"

    def _diff_findings(
        self,
        original_issues: List[Dict],
        patched_issues: List[Dict],
        hunks: List[PatchHunk],
    ) -> List[PatchFinding]:
        """Identify issues that are NEW in the patched version.

        An issue is considered 'new' if it:
        1. Falls within or near a hunk's modified line range, AND
        2. Was NOT present in the original (by fingerprint).

        This ensures we ONLY report issues that:
        (a) are in regions actually touched by the patch, and
        (b) did not already exist before the patch was applied.
        """
        # Build fingerprint set from original
        orig_fingerprints = {self._fingerprint_issue(i) for i in original_issues}

        # Build a set of line ranges modified by hunks (in the NEW/patched file)
        modified_ranges: List[Tuple[int, int]] = []
        for hunk in hunks:
            start = hunk.new_start
            end = start + hunk.new_count
            modified_ranges.append((start, end))

        def _in_modified_range(line: int) -> bool:
            """Check if a line falls within or near a modified hunk range."""
            for start, end in modified_ranges:
                if (start - 3) <= line <= (end + 3):
                    return True
            return False

        new_findings: List[PatchFinding] = []

        for issue in patched_issues:
            fp = self._fingerprint_issue(issue)
            line_num = issue.get("line_number", 0)

            # GATE 1: Issue MUST be in or near a modified hunk range.
            # This is the primary filter — ignore everything outside
            # the patch region regardless of fingerprint matching.
            if not _in_modified_range(line_num):
                continue

            # GATE 2: Issue must NOT have existed in the original file.
            if fp in orig_fingerprints:
                continue

            finding = PatchFinding(
                file_path=issue.get("file_path", self.filename),
                line_number=line_num,
                title=issue.get("title", ""),
                severity=issue.get("severity", "medium"),
                confidence=issue.get("confidence", ""),
                category=issue.get("category", ""),
                description=issue.get("description", ""),
                suggestion=issue.get("suggestion", ""),
                code_before=issue.get("code", ""),
                code_after=issue.get("fixed_code", ""),
                introduced_by_patch=True,
                issue_source=issue.get("source", "patch"),
            )
            new_findings.append(finding)

        return new_findings

    # ------------------------------------------------------------------
    # Excel output
    # ------------------------------------------------------------------

    def _update_excel(
        self, excel_path: str, findings: List[PatchFinding],
        adapter_results: Optional[Dict[str, List[Dict]]] = None,
    ) -> None:
        """Write patch findings to a ``patch_<filename>`` tab in the Excel file.

        The column layout matches the main ``Analysis`` sheet produced by
        :class:`CodebaseLLMAgent` so the fixer agent can consume it:

            S.No | Title | Severity | Confidence | Category | File | Line |
            Description | Suggestion | Code | Fixed_Code | Feedback | Constraints
        """
        if not EXCEL_WRITER_AVAILABLE:
            self.logger.warning("ExcelWriter not available — writing findings as JSON instead")
            self._write_findings_json(findings)
            return

        sheet_name = f"patch_{self.filename_stem}"
        # Truncate sheet name to Excel's 31-char limit
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        # Column layout matching CodebaseLLMAgent's Analysis sheet
        headers = [
            "S.No", "Title", "Severity", "Confidence", "Category",
            "File", "Line", "Description", "Suggestion",
            "Code", "Fixed_Code", "Feedback", "Constraints",
        ]

        data_rows: List[List[Any]] = []
        for idx, f in enumerate(findings, start=1):
            data_rows.append([
                idx,
                f.title or f.description[:80] if f.description else "",
                f.severity,
                f.confidence,
                f.category,
                f.file_path,
                f.line_number,
                f.description,
                f.suggestion,
                f.code_before,
                f.code_after,
                f.feedback,
                f.constraints,
            ])

        try:
            excel_file = Path(excel_path)

            if excel_file.exists():
                try:
                    import openpyxl
                    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                    wb = openpyxl.load_workbook(str(excel_file))

                    # Remove existing sheet with the same name if present
                    if sheet_name in wb.sheetnames:
                        del wb[sheet_name]

                    ws = wb.create_sheet(title=sheet_name)

                    # ── Style definitions (matching ExcelWriter defaults) ──
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill("solid", fgColor="4F81BD")
                    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    thin = Side(border_style="thin", color="D0D0D0")
                    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    alt_fill = PatternFill("solid", fgColor="F3F3F3")

                    # Severity colour fills
                    sev_fills = {
                        "CRITICAL": PatternFill("solid", fgColor="FFC7CE"),
                        "MEDIUM":   PatternFill("solid", fgColor="FFEB9C"),
                        "LOW":      PatternFill("solid", fgColor="C6EFCE"),
                    }
                    sev_fonts = {
                        "CRITICAL": Font(bold=True, color="9C0006"),
                        "MEDIUM":   Font(bold=True, color="9C6500"),
                        "LOW":      Font(bold=True, color="006100"),
                    }

                    # Write styled header row
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align

                    # Write data rows with formatting
                    sev_col_idx = headers.index("Severity")  # 0-based
                    for row_idx, row_data in enumerate(data_rows, 2):
                        for col_idx, value in enumerate(row_data, 1):
                            cell = ws.cell(row=row_idx, column=col_idx, value=value)
                            cell.border = cell_border
                            cell.alignment = Alignment(vertical="top", wrap_text=True)

                            # Severity column colouring
                            if (col_idx - 1) == sev_col_idx:
                                upper_val = str(value).strip().upper()
                                if upper_val in sev_fills:
                                    cell.fill = sev_fills[upper_val]
                                    cell.font = sev_fonts[upper_val]
                            elif row_idx % 2 == 0:
                                cell.fill = alt_fill

                    # Auto-fit column widths
                    for col_idx, header in enumerate(headers, 1):
                        max_len = len(header)
                        for row_data in data_rows:
                            if (col_idx - 1) < len(row_data):
                                val_len = len(str(row_data[col_idx - 1]))
                                max_len = max(max_len, min(val_len, 60))
                        ws.column_dimensions[
                            openpyxl.utils.get_column_letter(col_idx)
                        ].width = max_len + 4

                    # Freeze header row and add auto-filter
                    ws.freeze_panes = "A2"
                    last_col = openpyxl.utils.get_column_letter(len(headers))
                    ws.auto_filter.ref = f"A1:{last_col}{len(data_rows) + 1}"

                    # Write separate static adapter tabs if adapters were run
                    self._write_adapter_tabs_openpyxl(wb, adapter_results)

                    wb.save(str(excel_file))
                    self.logger.info(
                        f"Updated {excel_path} with '{sheet_name}' tab ({len(findings)} findings)"
                    )
                    return

                except PermissionError:
                    self.logger.error(f"Permission denied writing to {excel_path}. File may be open.")
                    self._write_findings_json(findings)
                    return
                except Exception as exc:
                    self.logger.warning(f"Failed to update existing Excel: {exc} — attempting create new")

            # Create new workbook with ExcelWriter (matches other sheets' styling)
            writer = ExcelWriter(str(excel_path))
            writer.add_table_sheet(
                headers=headers,
                data_rows=data_rows,
                sheet_name=sheet_name,
                status_column="Severity",
            )

            # Append separate static adapter tabs
            if adapter_results:
                adapter_headers = [
                    "File", "Function", "Line", "Description",
                    "Severity", "Category", "CWE",
                ]
                for adapter_name, details in adapter_results.items():
                    if not details:
                        continue
                    tab_name = f"patch_static_{adapter_name}"[:31]
                    rows = [
                        [
                            d.get("file", ""),
                            d.get("function", ""),
                            d.get("line", ""),
                            d.get("description", ""),
                            d.get("severity", ""),
                            d.get("category", ""),
                            d.get("cwe", ""),
                        ]
                        for d in details
                    ]
                    writer.add_table_sheet(
                        adapter_headers, rows,
                        tab_name,
                        status_column="Severity",
                    )

            writer.save()
            self.logger.info(
                f"Created {excel_path} with '{sheet_name}' tab ({len(findings)} findings)"
            )

        except Exception as exc:
            self.logger.error(f"Failed to write Excel: {exc}")
            self._write_findings_json(findings)

    def _write_adapter_tabs_openpyxl(
        self, wb: Any, adapter_results: Optional[Dict[str, List[Dict]]]
    ) -> None:
        """Write per-adapter static analysis tabs to an existing openpyxl workbook.

        Matches the ``static_<adapter>`` tab format from :class:`CodebaseLLMAgent`.
        """
        if not adapter_results:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            adapter_headers = [
                "File", "Function", "Line", "Description",
                "Severity", "Category", "CWE",
            ]

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin = Side(border_style="thin", color="D0D0D0")
            cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            alt_fill = PatternFill("solid", fgColor="F3F3F3")

            for adapter_name, details in adapter_results.items():
                if not details:
                    continue
                tab_name = f"patch_static_{adapter_name}"[:31]

                # Remove existing tab
                if tab_name in wb.sheetnames:
                    del wb[tab_name]

                ws = wb.create_sheet(title=tab_name)

                # Header row
                for col_idx, hdr in enumerate(adapter_headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=hdr)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align

                # Data rows
                for row_idx, d in enumerate(details, 2):
                    row_data = [
                        d.get("file", ""),
                        d.get("function", ""),
                        d.get("line", ""),
                        d.get("description", ""),
                        d.get("severity", ""),
                        d.get("category", ""),
                        d.get("cwe", ""),
                    ]
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = cell_border
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                        if row_idx % 2 == 0:
                            cell.fill = alt_fill

                # Auto-fit + freeze
                for col_idx, hdr in enumerate(adapter_headers, 1):
                    max_len = len(hdr)
                    for d in details:
                        keys = ["file", "function", "line", "description", "severity", "category", "cwe"]
                        if (col_idx - 1) < len(keys):
                            val_len = len(str(d.get(keys[col_idx - 1], "")))
                            max_len = max(max_len, min(val_len, 60))
                    ws.column_dimensions[
                        openpyxl.utils.get_column_letter(col_idx)
                    ].width = max_len + 4
                ws.freeze_panes = "A2"

                self.logger.info(
                    f"  Added '{tab_name}' tab ({len(details)} findings)"
                )
        except Exception as exc:
            self.logger.warning(f"Failed to write adapter tabs: {exc}")

    def _write_findings_json(self, findings: List[PatchFinding]) -> None:
        """Fallback: write findings as JSON."""
        import json

        json_path = self.output_dir / f"patch_{self.filename_stem}_findings.json"
        data = [self._finding_to_dict(f) for f in findings]
        try:
            with open(json_path, "w", encoding="utf-8") as fp:
                json.dump(data, fp, indent=2, default=str)
            self.logger.info(f"Findings written to {json_path}")
        except Exception as e:
            self.logger.error(f"Failed to write JSON fallback: {e}")

    @staticmethod
    def _finding_to_dict(finding: PatchFinding) -> Dict[str, Any]:
        """Convert a PatchFinding to a plain dict."""
        return {
            "file_path": finding.file_path,
            "line_number": finding.line_number,
            "title": finding.title,
            "severity": finding.severity,
            "confidence": finding.confidence,
            "category": finding.category,
            "description": finding.description,
            "suggestion": finding.suggestion,
            "code": finding.code_before,
            "fixed_code": finding.code_after,
            "feedback": finding.feedback,
            "constraints": finding.constraints,
            "introduced_by_patch": finding.introduced_by_patch,
            "issue_source": finding.issue_source,
        }

    # ------------------------------------------------------------------
    # Utilities
    # ------------------------------------------------------------------

    def get_patch_summary(self) -> Dict[str, Any]:
        """Parse the patch and return a summary without running full analysis."""
        if not self.patch_file.exists():
            return {"error": "Patch file not found"}

        try:
            patch_content = self.patch_file.read_text(encoding="utf-8", errors="replace")
            hunks = self._parse_patch(patch_content)

            total_added = sum(len(h.added_lines) for h in hunks)
            total_removed = sum(len(h.removed_lines) for h in hunks)

            return {
                "patch_file": str(self.patch_file),
                "target_file": str(self.file_path),
                "hunk_count": len(hunks),
                "lines_added": total_added,
                "lines_removed": total_removed,
                "net_change": total_added - total_removed,
                "hunks": [
                    {
                        "header": h.header,
                        "orig_range": f"{h.orig_start},{h.orig_count}",
                        "new_range": f"{h.new_start},{h.new_count}",
                        "added": len(h.added_lines),
                        "removed": len(h.removed_lines),
                    }
                    for h in hunks
                ],
            }
        except Exception as e:
             self.logger.error(f"Failed to generate patch summary: {e}")
             return {"error": str(e)}


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="CURE Codebase Patch Agent — Analyse patches for introduced issues",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )

    parser.add_argument(
        "--file-path",
        required=True,
        help="Path to the original source file",
    )
    parser.add_argument(
        "--patch-file",
        required=True,
        help="Path to the .patch/.diff file (unified diff format)",
    )
    parser.add_argument(
        "--excel-path",
        default=None,
        help="Path to detailed_code_review.xlsx to update",
    )
    parser.add_argument(
        "-d", "--out-dir",
        default="./out",
        help="Output directory",
    )
    parser.add_argument(
        "--config-file",
        default=None,
        help="Path to global_config.yaml",
    )
    parser.add_argument(
        "--llm-model",
        default=None,
        help="LLM model in provider::model format",
    )
    parser.add_argument(
        "--enable-adapters",
        action="store_true",
        default=False,
        help="Run deep static analysis adapters on patched file",
    )
    parser.add_argument(
        "--constraints-dir",
        default="agents/constraints",
        help="Directory containing constraint files",
    )
    parser.add_argument(
        "--exclude-dirs",
        default="",
        help="Comma-separated list of directories to exclude",
    )
    parser.add_argument(
        "--exclude-globs",
        default="",
        help="Comma-separated glob patterns to exclude",
    )
    parser.add_argument(
        "--include-custom-constraints",
        nargs="*",
        default=[],
        help="Additional constraint .md files to load",
    )
    parser.add_argument(
        "--codebase-path",
        default=None,
        help="Root of the codebase for header/context/call-stack resolution "
             "(defaults to parent directory of --file-path)",
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        default=False,
        help="Enable verbose logging",
    )

    args = parser.parse_args()

    # Load config
    config = None
    if GLOBAL_CONFIG_AVAILABLE and args.config_file:
        try:
            config = GlobalConfig(args.config_file)
        except Exception as e:
            print(f"WARNING: Could not load config: {e}")

    # Setup LLM tools
    llm_tools = None
    if LLM_TOOLS_AVAILABLE:
        try:
            if args.llm_model:
                llm_tools = LLMTools(model=args.llm_model)
            elif config:
                model_str = config.get("llm.model")
                llm_tools = LLMTools(model=model_str) if model_str else LLMTools()
            else:
                llm_tools = LLMTools()
        except Exception as e:
            print(f"WARNING: Could not initialise LLMTools: {e}")

    # Run agent
    exclude_dirs = [d.strip() for d in args.exclude_dirs.split(",") if d.strip()]
    exclude_globs = [g.strip() for g in args.exclude_globs.split(",") if g.strip()]

    agent = CodebasePatchAgent(
        file_path=args.file_path,
        patch_file=args.patch_file,
        output_dir=args.out_dir,
        config=config,
        llm_tools=llm_tools,
        enable_adapters=args.enable_adapters,
        verbose=args.verbose,
        constraints_dir=args.constraints_dir,
        exclude_dirs=exclude_dirs,
        exclude_globs=exclude_globs,
        custom_constraints=args.include_custom_constraints,
        codebase_path=args.codebase_path,
    )

    result = agent.run_analysis(excel_path=args.excel_path)

    print(f"\n{'='*60}")
    print(f" Patch Analysis Results: {agent.filename}")
    print(f"{'='*60}")
    print(f"  Status:           {result.get('status')}")
    print(f"  Hunks parsed:     {result.get('hunks_parsed', 0)}")
    print(f"  Original issues:  {result.get('original_issue_count', 0)}")
    print(f"  Patched issues:   {result.get('patched_issue_count', 0)}")
    print(f"  NEW issues:       {result.get('new_issue_count', 0)}")
    print(f"  Excel output:     {result.get('excel_path', 'N/A')}")
    print(f"{'='*60}")

    if result.get("findings"):
        print(f"\n  Findings:")
        for i, f in enumerate(result["findings"], 1):
            print(f"    {i}. [{f['severity']}] {f['category']} — {f['description'][:80]}")
            print(f"       Line {f['line_number']} in {f['file_path']}")

    sys.exit(0 if result.get("status") == "success" else 1)